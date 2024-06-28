import random
import string
import datetime
from openpyxl import Workbook

date = datetime.date.today().strftime("%d-%m-%Y")
wb = Workbook()
ws = wb.active
ws.title = "Coupon"

#######################################################################################################################

# TN = "TNTWSAS"
# KA = "KATWSAS"

# ws.cell(row=1,column=1).value = "TN"
# ws.cell(row=1,column=2).value = "KA"
#
# ws.cell(row=2,column=1).value = "TWS R100 - Black"
# ws.cell(row=2,column=2).value = "TWS R100 - Black"
#
# ws.cell(row=3,column=1).value = "M-True Wireless (TWS)"
# ws.cell(row=3,column=2).value = "M-True Wireless (TWS)"
#
# ws.cell(row=4,column=1).value = "Anker SoundCore"
# ws.cell(row=4,column=2).value = "Anker SoundCore"

# r = 1

# while range:
#     tn = TN + str(random.randint(1, 99999999))
#     ka = KA + str(random.randint(1, 99999999))
#     # print(len(tn))
#     if len(tn) == 14 and len(ka) == 14:
#         r = r + 1
#         if r == 600000+1:
#             break
#         print(r)
#         ws.cell(row=r, column=1).value = tn
#         ws.cell(row=r, column=2).value = ka

# for r in range(6,600000+1):
#     print(r)
#
#     ws.cell(row=r, column=1).value = TN + str(random.randint(1, 99999999))
#     ws.cell(row=r, column=2).value = KA + str(random.randint(1, 99999999))
# wb.save(r"D:\Durai\Coupon Code\Save Data\Coupon Code 1 "+date+".xlsx")

#######################################################################################################################
# r = 1
# while range:
#     coupon = str(random.randint(1, 999999999))
#     print(r)
#
#     if len(coupon) == 8:
#         ws.cell(row=r, column=1).value = coupon
#         r = r + 1
#
#     if r == 700:
#         break
#
# wb.save(r"D:\Durai\Coupon Code\Save Data\Coupon Code 1 " + date + ".xlsx")

#######################################################################################################################

N = 4
for r in range(1,191):
    coupon_1 = ''.join(random.choice(string.ascii_uppercase) for r1 in range(N))
    coupon_2 =''.join (random.choice(string.digits) for r2 in range(N))

    coupon = coupon_1 + coupon_2
    print(str(coupon))
    ws.cell(row=r, column=1).value = coupon

    wb.save(r"D:\Durai\Coupon Code\Save Data\Coupon Code " + date + ".xlsx")
