from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import time
import datetime
ac_wb = load_workbook(r"D:\Durai\Scraping\Home_appliances\Web Url\Urls.xlsx")
ac_ws = ac_wb.active
print("No Of Rows",ac_ws.max_row)
# print("No Of Columns",ac_ws.max_column)
# print(ac_ws.value)
date = datetime.date.today().strftime("%#d-%#m-%Y")



options = webdriver.ChromeOptions()
# prefs = {'profile.default_content_setting_values': {'cookies': 2, 'images': 2,  'popups': 2, 'geolocation':2,
prefs = {'profile.default_content_setting_values': {'notifications': 2, 'auto_select_certificate': 2, 'fullscreen': 2,'popups': 2,
                                                    'mouselock': 2, 'mixed_script': 2, 'media_stream': 2,'plugins': 2,
                                                    'media_stream_mic': 2, 'media_stream_camera': 2,
                                                    'protocol_handlers': 2,
                                                    'ppapi_broker': 2, 'automatic_downloads': 2, 'midi_sysex': 2,
                                                    'push_messaging': 2, 'ssl_cert_decisions': 2,
                                                    'metro_switch_to_desktop': 2,
                                                    'protected_media_identifier': 2, 'app_banner': 2,
                                                    'site_engagement': 2,
                                                    'durable_storage': 2}}
options.add_experimental_option('prefs', prefs)
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe", options=options)

save_wb = Workbook()
save_ws = save_wb.active

save_ws.cell(row=1, column=1).value = "Product Id"
save_ws.cell(row=1, column=2).value = "Item Code"
save_ws.cell(row=1, column=3).value = "Model"
save_ws.cell(row=1, column=4).value = "Poorvika"
save_ws.cell(row=1, column=5).value = "Amazon Price"
save_ws.cell(row=1, column=6).value = "Amazon Url"

for r in range(2, ac_ws.max_row+1):

    save_ws.cell(row=r, column=1).value = ac_ws.cell(row=r, column=1).value
    save_ws.cell(row=r, column=2).value = ac_ws.cell(row=r, column=2).value
    save_ws.cell(row=r, column=3).value = ac_ws.cell(row=r, column=3).value

    if ac_ws.cell(row=r, column=9).value != "NA":

        save_ws.cell(row=r, column=6).value = ac_ws.cell(row=r, column=9).value
        print("#" * 150)
        print("Amazon")
        print('Range: ', r)
        print("Amazon", ac_ws.cell(row=r, column=9).value)

        try:
            driver.get(url=ac_ws.cell(row=r, column=9).value)
            time.sleep(3)

            try:

                for price1 in driver.find_elements(By.ID, "olp_feature_div"):
                    for price2 in price1.find_elements(By.CLASS_NAME, "a-size-base"):
                        print("Amazon Price 3 = ", price2.text[1:])
                        save_ws.cell(row=r, column=5).value = price2.text[1:]
            except:
                pass

            try:

                price = driver.find_element(By.ID, "apex_desktop")
                price1 = price.find_element(By.CLASS_NAME, "a-price-whole")
                print("Amazon Price 2 = ", price1.text)
                save_ws.cell(row=r, column=5).value = price1.text
            except:
                pass

            try:

                price = driver.find_element(By.ID, "apex_desktop")
                price2 = price.find_element(By.CLASS_NAME, "apexPriceToPay")
                print("Amazon Price 1 = ", price2.text[1:])
                save_ws.cell(row=r, column=5).value = price2.text[1:]
            except:
                pass

        except:
            pass

    save_wb.save(r"D:\Durai\Scraping\Home_appliances\Save Date's\Save Files\Total Scraping\Home Application Amazon " + date + ".xlsx")

driver.quit()
