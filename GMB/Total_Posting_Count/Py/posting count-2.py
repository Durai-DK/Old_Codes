from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import datetime
from GMB.Google.Google_login import Google


date = datetime.datetime.now().strftime("%d-%m-%Y")
wb = Workbook()
ws = wb.active
driver = webdriver.Chrome(executable_path=r"D:/Durai/Driver/chromedriver.exe")
Gg = Google(driver=driver)
Gg.login()

post_wb = load_workbook(r"D:/Durai/GMB/Posting/Data/Posting post urls.xlsx")
post_ws = post_wb.active


time.sleep(5)
driver.maximize_window()
driver.implicitly_wait(30)
l = 1
for r in range(285, 300):
    print("")
    print("#" * 150)
    print("")
    print(r)
    driver.get(post_ws.cell(row=r, column=2).value)
    driver.implicitly_wait(5)

    ws.cell(row=l, column=1).value = driver.find_element(By.CLASS_NAME,"My5R8c").text

    for data in driver.find_elements(By.CLASS_NAME,"p4HiUc"):
        # print(data.text)
        for head in data.find_elements_by_class_name("LgQiCc"):
            for name in head.find_elements_by_class_name("P9ZBeb"):
                print(name.text)
                ws.cell(row=l,column=2).value = name.text

            for view in head.find_elements_by_class_name("G5bhdb"):
                if view.text[-5:] == "views":
                    print(view.text)
                    ws.cell(row=l, column=3).value = view.text
                if view.text[-6:] == "clicks" or view.text[-5:] == "click":
                    ws.cell(row=l, column=4).value = view.text
                    print(view.text)

            for date_value in head.find_elements_by_class_name("PROnRd"):
                print(date_value.text)
                ws.cell(row=l,column=5).value = date_value.text
                l = l + 1
                wb.save(r"D:\Durai\GMB\Total_Posting_Count\Save Data\count 285-300 "+ date +".xlsx")
driver.quit()

# 462
# 406
# 284

