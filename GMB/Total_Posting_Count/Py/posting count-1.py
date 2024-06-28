from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import datetime
from GMB.Google.Google_login import Google

date = datetime.datetime.now().strftime("%d-%m-%Y")

wb = Workbook()
ws = wb.active

driver = webdriver.Chrome(executable_path=r"D:/Durai/Driver/chromedriver.exe")
Gg = Google(driver=driver)
Gg.login()

post_wb = load_workbook(r"D:/Durai/GMB/Posting/Data/Posting post urls.xlsx")
post_ws = post_wb.active

time.sleep(5)
driver.maximize_window()
driver.implicitly_wait(30)
l = 1
for r in range(2, 5):
    print("")
    print(r)
    print("link = ",post_ws.cell(row=r, column=2).value)

    driver.get(post_ws.cell(row=r, column=2).value)
    time.sleep(3)


    for name in driver.find_elements(By.ID, "lb"):
        # driver.execute_script("window.scrollBy(0,2000)", "")
        print("done")

        for name_1 in name.find_elements(By.TAG_NAME, "undefined"):
            print(name_1.text)


