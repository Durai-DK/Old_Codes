import time
import datetime
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import load_workbook,Workbook
from Google.Google_login import Google

driver = webdriver.Chrome(service=Service("D:\Durai\Driver\chromedriver.exe"))
date = datetime.date.today().strftime("%d/%m/%Y")
Date = datetime.date.today().strftime("%d-%m-%Y")


# wb = load_workbook(r"D:\Durai\GMB\Posting\Data\Posting post urls.xlsx")
wb = load_workbook(r"D:\Durai\GMB\Posting\Data\TamilNadu Post URL.xlsx")
# wb = load_workbook(r"D:\Durai\GMB\Posting\Data\Posting post urls - Copy.xlsx")
ws = wb.active

fb = Workbook()
fs = fb.active

d_b = Workbook()
d_s = d_b.active

class GooglePosting:

    def move_to_frame(self):
        time.sleep(1)
        for button in driver.find_elements(By.CLASS_NAME, 'tN4Gf'):
            if button.get_attribute('jsname') == 'z47ex':
                button.click()

        for event in driver.find_elements(By.TAG_NAME, 'iframe'):
            if event.get_attribute('sandbox') == 'allow-same-origin allow-scripts allow-forms allow-popups allow-popups-to-escape-sandbox':
                driver.switch_to.frame(event)
                # print('Switched')
                driver.implicitly_wait(5)

    def post_title(self, **kwargs):
        time.sleep(1)
        desc = driver.find_element(By.ID, 'i6')
        desc.send_keys(kwargs.get('title'))
        print("Titled")
        driver.implicitly_wait(1)

    def image_upload_form(self):
        time.sleep(1)
        for upload in driver.find_elements(By.CLASS_NAME, 'hqTBzf'):
            buttonclick = upload.find_element(By.TAG_NAME, 'button')
            buttonclick.click()
            driver.implicitly_wait(15)
            time.sleep(1)


    def image_upload(self, **kwargs):
        time.sleep(1)
        test3 = driver.find_element(By.CLASS_NAME, 'picker-dialog-content')
        driver.switch_to.frame(test3.find_element(By.TAG_NAME, 'iframe'))
        driver.implicitly_wait(5)
        # print("image")
        image = driver.find_element(By.TAG_NAME, 'input')
        image.send_keys(kwargs.get('image'))
        time.sleep(3)
        print("image uploaded")
        driver.implicitly_wait(10)
        driver.switch_to.default_content()
        for event in driver.find_elements(By.TAG_NAME, 'iframe'):
            if event.get_attribute(
                    'sandbox') == 'allow-same-origin allow-scripts allow-forms allow-popups allow-popups-to-escape-sandbox':
                driver.switch_to.frame(event)
                driver.implicitly_wait(5)
        driver.implicitly_wait(5)

    def end_date(self, **kwargs):
        time.sleep(1)

        driver.find_element(By.ID, 'i11').send_keys(date)
        # End date
        driver.find_element(By.ID, 'i22').send_keys(kwargs.get('date'))
        print("Date Updated")
        time.sleep(1)

    def post_description(self, **kwargs):
        time.sleep(1)
        driver.find_element(By.ID, 'i31').send_keys(kwargs.get('description'))
        time.sleep(2)

    def post_published(self):
        time.sleep(1)
        posts = driver.find_element(By.CLASS_NAME, 'FkJOzc')
        for clicked in posts.find_elements(By.TAG_NAME, 'button'):
            post = clicked.get_attribute('jsname')
            # print(post)
            if post == "vdQQuc":
                clicked.click()
                print("Successfully Posted")

    def post_call(self, **kwargs):
        time.sleep(1)
        button = driver.find_element(By.CLASS_NAME, 'VfPpkd-LgbsSe')
        button.find_element(By.CLASS_NAME, 'VfPpkd-kBDsod').click()

        for field in driver.find_elements(By.CLASS_NAME, 'VfPpkd-StrnGf-rymPhb-b9t22c'):
            if field.text == kwargs.get('field'):
                print(kwargs.get('field'))
                field.click()

class GooglePostingRun:

    def __init__(self, posting=None):
        self.posting = posting

    def range_run(self, **kwargs):
        google = Google(driver=driver)
        google.login()
        time.sleep(5)

        print(self.posting.Title)
        driver.implicitly_wait(3)

        for num in range(kwargs.get('start'), kwargs.get('end')+1):
            print(num)
            fs.cell(row=1, column=1).value = num
            fb.save(r"D:\Durai\GMB\Posting\Posting_lastPost\findlastpost " + str(kwargs.get('value')) + ".xlsx")
            print(ws.cell(row=num, column=2).value)

            try:
                driver.get(ws.cell(row=num, column=2).value)
                driver.implicitly_wait(3)
                # time.sleep(10)
                self.posted()

            except:

                d_s.cell(row=num, column=1).value = num

                d_b.save(r"D:\Durai\GMB\Posting\Posting_lastPost\post skip numbers " + str(kwargs.get('value')) + " "+Date + ".xlsx")



        driver.close()
        driver.quit()

    def posted(self):
        time.sleep(2)
        gp = GooglePosting()
        gp.move_to_frame()
        gp.post_title(title=self.posting.Title)
        gp.image_upload_form()
        gp.image_upload(image=self.posting.Img_file)
        gp.end_date(date=self.posting.End_date)
        gp.post_description(description=self.posting.Description)
        gp.post_call(field=self.posting.Field)
        # time.sleep(5)
        gp.post_published()

