from Posting.Posting_File.posting_forms import *
from Posting.Posting_Fields.posting_fields import *
from openpyxl import load_workbook


wb = load_workbook(r'D:\Durai\GMB\Posting\Posting_lastPost\findlastpost 5.xlsx')
ws = wb.active


g1 = GooglePostingRun(posting=PostingField3)
g1.range_run(start=ws.cell(row=1,column=1).value, end=460,value=1)