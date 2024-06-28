from selenium import webdriver
from openpyxl import load_workbook,Workbook
from selenium.webdriver.common.by import By
from GMB.Google.Google_login import Google
import time
from GMB.product.Product_Field.product_fields import ProductField
"None"
"Order online"
"Buy"
"Learn more"
"Get offer"

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")
wb = load_workbook(r"D:\Durai\GMB\product\Data\GMB Product URL.xlsx")
ws = wb.active

d_wb = Workbook()
d_ws = d_wb.active

class GoogleProduct:
    def __init__(self):
        self.driver = driver

    def product_add(self):
        for clic in driver.find_elements(By.CLASS_NAME, "hHErVc"):
            clic.find_element(By.CLASS_NAME, "VfPpkd-LgbsSe").click()
        time.sleep(3)

    def product_image(self):
        self.driver.implicitly_wait(10)
        self.driver.find_element(By.CLASS_NAME, "ULhCdf").send_keys(ProductField.Image)
        time.sleep(3)

    def product_title(self):
        try:
            self.driver.find_element(By.ID, 'c3').send_keys(ProductField.Title)
            self.driver.implicitly_wait(10)

        except:
            self.driver.find_element(By.CLASS_NAME, 'ZUPIVd').send_keys(ProductField.Title)
            self.driver.implicitly_wait(10)

    def product_category(self):
        self.driver.find_element(By.CLASS_NAME, "VfPpkd-TkwUic").click()
        time.sleep(5)

        try:
            for r2 in self.driver.find_elements(By.CLASS_NAME, 'VfPpkd-StrnGf-rymPhb-ibnC6b'):
                if r2.text == ProductField.Category:
                    r2.click()
                elif r2.get_attribute("data-value") == "new":
                    r2.click()
                    time.sleep(2)
                    self.driver.find_element(By.ID, 'c11').send_keys(ProductField.Category)

        except:
            self.driver.find_element(By.ID, 'c11').send_keys(ProductField.Category)

    def product_description(self):
        self.driver.find_element(By.ID, 'c27').send_keys(ProductField.Description)
        self.driver.implicitly_wait(10)

    def product_optional_button(self):
        for dr in self.driver.find_elements(By.CLASS_NAME, "WEGjDf"):
            if dr.text == "Add a button (optional)":
                dr.click()
        time.sleep(3)
        for r1 in self.driver.find_elements(By.CLASS_NAME, "VfPpkd-StrnGf-rymPhb-b9t22c"):
            if r1.text == ProductField.Field:
                r1.click()

        self.driver.find_element(By.ID, "c35").send_keys(ProductField.Online_Link)

    def product_special_click(self):
        self.driver.find_element(By.ID, "c36").click()

    def product_save(self):
        for r1 in self.driver.find_elements(By.CLASS_NAME, "VfPpkd-vQzf8d"):
            if r1.text == "Save":
                print("Done Save")
                r1.click()


class GoogleProductRun:
    def __init__(self):
        pass

    def range_run(self,**kwargs):

        G = Google(driver=driver)
        G.login()
        time.sleep(2)
        for num in range(kwargs.get('start'),kwargs.get('end')+1):
            print(num)
            time.sleep(7)
            print(ws.cell(row=num,column=2).value)

            d_ws.cell(row=1,column=1).value = num
            d_wb.save(r"D:\Durai\GMB\product\Product load\Product page info " + str(kwargs.get("value")) + ".xlsx")
            driver.get(url=ws.cell(row=num,column=2).value)


            driver.implicitly_wait(10)
            gp = GoogleProduct()
            gp.product_add()
            gp.product_image()
            gp.product_title()
            gp.product_category()
            gp.product_description()
            gp.product_optional_button()
            gp.product_special_click()
            gp.product_save()

        driver.quit()
        driver.close()