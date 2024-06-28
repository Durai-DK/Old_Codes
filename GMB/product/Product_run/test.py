from selenium import webdriver
from openpyxl import load_workbook,Workbook
from selenium.webdriver.common.by import By
from GMB.Google.Google_login import Google
import time
from GMB.product.Product_Field.product_fields import *
"None"
"Order online"
"Buy"
"Learn more"
"Get offer"

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")

wb = load_workbook(r"D:\Durai\GMB\product\Data\GMB Product URL.xlsx")
ws = wb.active

d_wb = Workbook()
d_ws = d_wb.active

class GoogleProduct:
    def __init__(self):
        self.driver = driver

    def product_add(self):

        for clic in driver.find_elements(By.CLASS_NAME, "hHErVc"):
            clic.find_element(By.CLASS_NAME, "VfPpkd-LgbsSe").click()
        time.sleep(3)

    def product_image(self, **kwargs):

        driver.implicitly_wait(10)
        image = driver.find_element(By.CLASS_NAME, "ULhCdf")
        image.send_keys(kwargs.get("image"))
        time.sleep(3)

    def product_title(self, **kwargs):
        try:
            title = driver.find_element(By.ID, 'c3')
            title.send_keys(kwargs.get("title"))
            driver.implicitly_wait(10)

        except:
            title = driver.find_element(By.CLASS_NAME, 'ZUPIVd')
            title.send_keys(kwargs.get("title"))
            driver.implicitly_wait(10)

    def product_category(self, **kwargs):
        driver.find_element(By.CLASS_NAME, "VfPpkd-TkwUic").click()
        time.sleep(5)

        try:
            for r2 in driver.find_elements(By.CLASS_NAME, 'VfPpkd-StrnGf-rymPhb-ibnC6b'):
                if r2.text == kwargs.get("Category"):
                    r2.click()
                elif r2.get_attribute("data-value") == "new":
                    r2.click()
                    time.sleep(2)
                    driver.find_element(By.ID, 'c11').send_keys(kwargs.get("Category"))

        except:
            driver.find_element(By.ID, 'c11').send_keys(kwargs.get("Category"))

    def product_description(self, **kwargs):
        driver.find_element(By.ID, 'c27').send_keys(kwargs.get("Description"))
        driver.implicitly_wait(10)

    def product_optional_button(self, **kwargs):
        for dr in driver.find_elements(By.CLASS_NAME, "WEGjDf"):
            if dr.text == "Add a button (optional)":
                dr.click()
        time.sleep(3)
        for r1 in driver.find_elements(By.CLASS_NAME, "VfPpkd-StrnGf-rymPhb-b9t22c"):
            if r1.text == kwargs.get("link"):
                r1.click()

        new = driver.find_element(By.ID, "c35")
        new.send_keys(kwargs.get("new"))

    def product_special_click(self):
        driver.find_element(By.ID, "c36").click()

    def product_save(self):
        for r1 in driver.find_elements(By.CLASS_NAME, "VfPpkd-vQzf8d"):
            if r1.text == "Save":
                print("Done Save")
                r1.click()


class GoogleProductRun:
    def __init__(self,product=None):
        self.product = product
        pass

    def range_run(self,**kwargs):
        Gg = Google(driver=driver)
        Gg.login()
        # url = "https://business.google.com/u/3/products/l/02083316379263937574"
        # driver.get(url)
        time.sleep(2)
        for num in range(kwargs.get('start'),kwargs.get('end')+1):
            print(num)
            time.sleep(7)
            print(ws.cell(row=num,column=2).value)

            d_ws.cell(row=1,column=1).value = num
            d_wb.save(r"D:\Durai\GMB\product\Product load\Product page info " + str(kwargs.get("value")) + ".xlsx")
            driver.get(url=ws.cell(row=num,column=2).value)

    def outdata(self):

        driver.implicitly_wait(10)
        gp = GoogleProduct()
        gp.product_add()
        gp.product_image(image=self.product.Image)
        gp.product_title(title=self.product.Title)
        gp.product_category(Category=self.product.Category)
        gp.product_description(Description=self.product.Description)
        gp.product_optional_button(link=self.product.Online_Link,new=self.product.Field)
        gp.product_special_click()
        # gp.product_save()

        driver.quit()
        driver.close()
