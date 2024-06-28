from GMB.product.form.product_forms import GoogleProductRun
from openpyxl import load_workbook


dk_wb = load_workbook(r"D:\Durai\GMB\product\Product load\Product page info 5.xlsx")
dk_ws = dk_wb.active


g = GoogleProductRun()
g.range_run(start=dk_ws.cell(row=1,column=1).value,end=470,value=5)