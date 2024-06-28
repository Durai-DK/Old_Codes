from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from GMB.Google.Google_login import Google
import time



wb = load_workbook(r"D:\Durai\GMB\product\Data\GMB Product URL.xlsx")
ws = wb.active

driver = webdriver.Chrome(r"D:\Durai\Driver\chromedriver.exe")
g = Google(driver=driver)
g.login()

data_list = ["Own Premium MacBooks at ₹10,000* Inst. Discou…",
        "Grab New Smartphones. Now @Upto ₹15,000* Ins…",
        "Get Upto ₹5,000* Cashback & Freebies wit…",
        "Take Home New Smart TVs. Upto 10%* Inst. Dis…",
        "Increase Productivity with Tablets & iPads @₹4,000…",
        "True Music Freedom. TWS Earbuds @₹2,500* Off",
        "Listen Anywhere, Anytime. Speakers @10%* Inst. Off",
        "Get the World at your Wrist@₹4,000* Inst.Off! …",
        "For Smarter Living. ₹1000* Inst. Off on Smar…",
        "Taking Care of Yourself Got Easier at Upto 10%* …",
        "Avoid Tangles with Wirless Headphones. 10…"]


class Run:
    def All_Product(self):
        for products in driver.find_elements(By.CLASS_NAME,'zTXfQc'):
            for product in products.find_elements(By.CLASS_NAME,'bgmvLc'):
                for r in range(0,len(data_list)):
                    if product.text in data_list[r]:
                        print(product.text)
                        Run.product_click(self)
                        Run.product_delete(self)
                        Run.confim_delete(self)
                        break

    def product_click(self):
        driver.find_element(By.CLASS_NAME, "hhfn6").click()
        time.sleep(1)



    def product_delete(self):
        driver.find_element(By.CLASS_NAME,"mN1ivc").click()


    def confim_delete(self):
        for delete_box in driver.find_elements(By.CLASS_NAME, "XfpsVe"):
            for del_in in delete_box.find_elements(By.CLASS_NAME, "RveJvd"):
                # print(del_in.text)
                if del_in.text == "Delete":
                    del_in.click()
                    print("done")
                    driver.implicitly_wait(5)
                    time.sleep(5)

    def Product(self):
        for r in range(197,456):
            print(r)
            time.sleep(3)
            print(ws.cell(row=r, column=2).value)
            driver.get(url=ws.cell(row=r, column=2).value)
            driver.implicitly_wait(5)
            Run.All_Product(self)


c = Run()
c.Product()