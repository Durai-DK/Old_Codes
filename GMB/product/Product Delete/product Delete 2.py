from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from GMB.Google.Google_login import Google
import time

wb = load_workbook(r"D:\Durai\GMB\product\Data\GMB Product URL.xlsx")
ws = wb.active

driver = webdriver.Chrome(r"D:\Durai\Driver\chromedriver.exe")
g = Google(driver=driver)
g.login()

l = ["Own Premium MacBooks at ₹10,000* Inst. Discou…",
     "Grab New Smartphones. Now @Upto ₹15,000* Ins…",
     "Get Upto ₹5,000* Cashback & Freebies wit…",
     "Take Home New Smart TVs. Upto 10%* Inst. Dis…",
     "Increase Productivity with Tablets & iPads @₹4,000…",
     "True Music Freedom. TWS Earbuds @₹2,500* Off",
     "Listen Anywhere, Anytime. Speakers @10%* Inst. Off",
     "Get the World at your Wrist@₹4,000* Inst.Off! …",
     "For Smarter Living. ₹1000* Inst. Off on Smar…",
     "Taking Care of Yourself Got Easier at Upto 10%* …",
     "Avoid Tangles with Wirless Headphones. 10…"]

for r in range(2,450):

    print("")
    print(r)
    time.sleep(5)
    print(ws.cell(row=r,column=2).value)

    try:
        driver.get(url=ws.cell(row=r,column=2).value)
        driver.implicitly_wait(5)
        for r1 in driver.find_elements(By.CLASS_NAME,'zTXfQc'):
            for r2 in r1.find_elements(By.CLASS_NAME,'VfPpkd-ksKsZd-XxIAqe'):
                for r3 in r2.find_elements(By.CLASS_NAME,"bgmvLc"):

                    for _ in range(0,12):
                        print("web link :",r3.text)
                        print("list :",l[_])
                        if r3.text == l[_]:
                            r2.click()
                            time.sleep(2)

                    # if r3.text == "You're Lucky! Get ₹4,000* Inst. Cashback on Smart…":
                    #     r2.click()
                    #     time.sleep(2)

                            for dk1 in driver.find_elements(By.CLASS_NAME, "l72iR"):
                                for dk2 in dk1.find_elements(By.CLASS_NAME, "VfPpkd-Bz112c-LgbsSe"):

                                    if dk2.text == "delete":
                                        dk2.click()

                                        for lk1 in driver.find_elements(By.CLASS_NAME, "XfpsVe"):
                                            for lk2 in lk1.find_elements(By.CLASS_NAME, "RveJvd"):
                                                # print(lk2.text)

                                                if lk2.text == "Delete":
                                                    lk2.click()
                                                    print("done")
                                                    break
    except:
            pass
driver.quit()
driver.close()
