from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from GMB.Google.Google_login import Google
'Create a new category'
'Mobile'

"Order online"
""
import time
wb = load_workbook(r"D:\Durai\GMB\product\Data\GMB Product URL.xlsx")
ws = wb.active

driver = webdriver.Chrome(r"D:\Durai\Driver\chromedriver.exe")
g = Google(driver=driver)
g.login()

image = r"D:\Durai\GMB\product\Image\17-3-2022-1.jpeg"
url = "https://business.google.com/u/3/products/l/02083316379263937574"
new_category = "Mobile"
Description = "Splash more awesomeness and productivity over your life. Get New A15 Bionic powered Apple iPhone 13 and Upgrade to a premium Apple lifestyle. Visit Poorvika Online or Walk-in to your nearest Poorvika Showroom and get benefits worth upto ₹25,000*! Instant Store Discount of upto ₹8,000*, Cashback upto ₹6000*, upto ₹3000* Exchange Bonus and More! Hurry Now! Valid till 20th March*."
Online_Order = 'https://bit.ly/Apple_iPhone_13'


for r in range(369,371):
    print(r)
    time.sleep(10)
    print(ws.cell(row=r,column=2).value)
    try:
        driver.get(url=ws.cell(row=r,column=2).value)
        driver.implicitly_wait(5)
        for r1 in driver.find_elements(By.CLASS_NAME,'zTXfQc'):
            for r2 in r1.find_elements(By.CLASS_NAME,'TB4jmc'):
                for r3 in r2.find_elements(By.CLASS_NAME,"bgmvLc"):
                    print(r3.text)
                    if r3.text == "Samsung Galaxy Z Flip 4 5G":
                        r2.click()
                        time.sleep(3)

                        driver.implicitly_wait(10)
                        driver.find_element(By.CLASS_NAME, "ULhCdf").send_keys(r"D:\Durai\GMB\product\image\26-08-2022-2.jpeg")
                        time.sleep(5)

                        for l1 in driver.find_elements(By.CLASS_NAME, "VfPpkd-vQzf8d"):
                            if l1.text == "Save":
                                print("Save")
                                l1.click()

    except:
        pass
driver.quit()
driver.close()
